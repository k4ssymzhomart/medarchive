"""Экстрактор XLSX (раздел 7.4).

Особенности реального архива:
  - Клиника 6: 5181 строка, в начале листа служебная преамбула приказа,
    строка заголовка НЕ первая, а где-то на 6-8 строке.
  - Клиника 8: два листа «Страховой» и «востребованные», 8-строчная шапка
    приказа.
  - Категории/секции часто оформлены объединёнными ячейками (merge_cells):
    строка только с текстом и без цены = заголовок раздела, тянем его вперёд.

Логика тупая и заменяемая: находим строку заголовка по якорным словам,
сопоставляем колонки по подписям, дальше читаем имя/код/цены через parse_amount.
Раскладка резидент/нерезидент происходит позже в price_parser.
"""

from __future__ import annotations

from decimal import Decimal

from openpyxl import load_workbook

from app.pipeline.base import ExtractedItem, ExtractionResult, Extractor
from app.pipeline.columns import HEADER_ANCHORS, NAME_HINTS, PRICE_HINTS
from app.pipeline.price_parser import parse_amount

# Сильные подписи колонки кода. Намеренно БЕЗ «№»: одинокая «№ п/п» — это
# колонка нумерации строк, а не код услуги (ловушка Клиники 6, где «№ п/п»
# перехватывала код у «Код услуги»).
_CODE_CAPTION_HINTS = ("код", "шифр", "тарификатор")

# Подписи тарифных колонок (помимо общих PRICE_HINTS): резидент/нерезидент,
# гражданство, страховка, первичный/повторный — всё это колонки с ценой.
# «стоимост» добавлено отдельно (PRICE_HINTS содержит «стоимость» целиком).
_PRICE_CAPTION_HINTS = PRICE_HINTS + (
    "резидент", "нерезидент", "граждан", "иностран", "снг",
    "зарубеж", "первичный", "повторный", "первичн", "повторн",
    "страхов", "стоимост",
)

# Стоп-слова строк-итогов: такие строки пропускаем целиком.
_TOTAL_WORDS = ("итого", "всего", "сумма прописью", "подытог", "subtotal", "total")

# Предельный размер текстового дампа raw_content (~500k символов).
_RAW_LIMIT = 500_000


class XlsxExtractor(Extractor):
    """Извлекает позиции из XLSX. Перебирает ВСЕ листы книги."""

    name = "xlsx"

    def can_handle(self, file_path: str, file_format: str) -> bool:
        return file_format == "xlsx"

    def extract(self, file_path: str) -> ExtractionResult:
        result = ExtractionResult(extractor_used=self.name)

        try:
            # read_only — экономия памяти на больших книгах (5181 строка);
            # data_only — берём посчитанные значения формул, а не сами формулы.
            wb = load_workbook(file_path, read_only=True, data_only=True)
        except Exception as exc:  # noqa: BLE001 — не падаем на битом файле
            result.warn(f"Не удалось открыть XLSX: {exc}")
            return result

        sheet_names = wb.sheetnames
        result.page_count = len(sheet_names)
        raw_parts: list[str] = []

        try:
            for sheet_idx, sheet_name in enumerate(wb.worksheets):
                try:
                    sheet_items, sheet_raw = self._extract_sheet(
                        sheet_name, sheet_idx, result
                    )
                except Exception as exc:  # noqa: BLE001 — лист не валит всю книгу
                    result.warn(
                        f"Лист «{sheet_name.title}» (#{sheet_idx}): ошибка {exc}"
                    )
                    continue

                if not sheet_items:
                    result.warn(
                        f"Лист «{sheet_name.title}» (#{sheet_idx}): "
                        "распознаваемых данных нет, пропущен"
                    )
                    continue

                result.items.extend(sheet_items)
                if sheet_raw:
                    raw_parts.append(f"# Лист «{sheet_name.title}»")
                    raw_parts.append(sheet_raw)
        finally:
            wb.close()

        result.raw_content = "\n".join(raw_parts)[:_RAW_LIMIT]
        return result

    # ------------------------------------------------------------------
    # Один лист
    # ------------------------------------------------------------------
    def _extract_sheet(
        self, sheet, sheet_idx: int, result: ExtractionResult
    ) -> tuple[list[ExtractedItem], str]:
        """Возвращает (позиции, текстовый дамп) для одного листа.

        sheet_idx — индекс листа (source_page). Excel-строки 1-based —
        source_row хранит реальный номер строки на листе.
        """
        # Материализуем строки как списки значений (read_only отдаёт генератор).
        rows: list[list] = []
        for row in sheet.iter_rows(values_only=True):
            rows.append(list(row) if row is not None else [])

        if not rows:
            return [], ""

        header_start, header_end, header_cells = self._find_header(rows)
        if header_start < 0:
            # Заголовок не нашли — данных на этом листе для нас нет.
            return [], ""

        col_map = self._map_columns(header_cells)
        if col_map["name"] is None:
            # Без колонки наименования распознавать нечего.
            return [], ""

        items: list[ExtractedItem] = []
        raw_lines: list[str] = []
        category: str | None = None

        # Данные идут после ВСЕЙ полосы заголовка (шапка бывает двухстрочной).
        for offset, row in enumerate(rows[header_end + 1:], start=1):
            excel_row = header_end + 1 + offset  # 1-based номер строки в Excel

            name = self._cell_text(self._cell_at(row, col_map["name"]))
            # Если в колонке имени пусто — пробуем найти любую текстовую ячейку
            # (объединённые секции часто «съезжают» в первую колонку).
            if not name:
                name = self._first_text(row)

            if not name:
                continue

            # Имя услуги обязано содержать буквы. Чисто числовые/символьные
            # «имена» — это строка нумерации колонок («1 2 3 4 5») сразу под
            # заголовком (Клиника 6) или мусор. Пропускаем.
            if not any(ch.isalpha() for ch in name):
                continue

            low = name.lower()
            # Пропускаем итоги/подытоги.
            if any(w in low for w in _TOTAL_WORDS):
                continue

            # Собираем цены по подписям колонок.
            prices: dict[str, Decimal] = {}
            for col_idx, label in col_map["price_labels"].items():
                amount = parse_amount(self._cell_at(row, col_idx))
                if amount is not None:
                    prices[label] = amount

            code = None
            if col_map["code"] is not None:
                code_text = self._cell_text(self._cell_at(row, col_map["code"]))
                code = code_text or None

            # Строка с текстом, но без единой цены = заголовок раздела/категории.
            # Запоминаем и тянем вперёд (merge_cells секций именно так).
            if not prices:
                # Эвристика: если нет ни цены, ни кода — это секция, не услуга.
                if code is None:
                    category = name
                    raw_lines.append(f"[категория] {name}")
                continue

            item = ExtractedItem(
                service_name_raw=name,
                service_code_source=code,
                prices=prices,
                category=category,
                source_page=sheet_idx,
                source_row=excel_row,
            )
            items.append(item)

            price_dump = " | ".join(f"{k}={v}" for k, v in prices.items())
            code_dump = f" [{code}]" if code else ""
            raw_lines.append(f"{excel_row}:{code_dump} {name} -> {price_dump}")

        return items, "\n".join(raw_lines)

    # ------------------------------------------------------------------
    # Поиск строки заголовка
    # ------------------------------------------------------------------
    def _find_header(self, rows: list[list]) -> tuple[int, int, list[str]]:
        """Ищет строку (или полосу строк) заголовка по якорным словам.

        Заголовок НЕ обязательно первый — у Клиник 6/7/8 сверху преамбула
        приказа на 6-8 строк. Более того, у Клиники 8 шапка двухстрочная:
        «Наименование услуги» и «Код по тарификатору» на одной строке, а
        подпись ценовой колонки «для граждан РК» — на соседней. Поэтому мы
        находим лучшую строку-якорь и сливаем её с соседями в одну полосу.

        Возвращает (индекс_начала, индекс_конца, слитые_ячейки) или (-1,-1,[]).
        Данные начинаются со строки индекс_конца + 1.
        """
        best_idx, best_score = -1, 0
        # Сканируем достаточно глубоко, чтобы перешагнуть преамбулу.
        for idx, row in enumerate(rows[:40]):
            joined = " ".join(self._cell_text(c) for c in row).lower()
            if not joined.strip():
                continue
            score = sum(1 for a in HEADER_ANCHORS if a in joined)
            if score > best_score:
                best_score, best_idx = score, idx

        if best_score < 2 or best_idx < 0:
            return -1, -1, []

        # Якорная строка задаёт структуру колонок. Соседние строки — это
        # продолжение шапки (двухстрочные заголовки Клиники 8), но они лишь
        # ДОПОЛНЯЮТ колонки, пустые в якорной строке (например подпись цены
        # «для граждан РК» под пустой ячейкой). Так мы не втягиваем строки
        # заголовка-титула («ПРЕЙСКУРАНТ», «цен на медицинские услуги»),
        # которые заняли бы уже заполненную колонку №/наименования.
        anchor = [self._cell_text(c) for c in rows[best_idx]]
        start = end = best_idx

        while start - 1 >= 0 and self._is_header_continuation(rows[start - 1], anchor):
            start -= 1
        while end + 1 < len(rows) and self._is_header_continuation(rows[end + 1], anchor):
            end += 1

        merged = self._merge_header_band(rows[start:end + 1], best_idx - start)
        return start, end, merged

    def _is_header_continuation(self, row: list, anchor: list[str]) -> bool:
        """Строка — продолжение шапки: несёт подпись-якорь, без цен, и
        заполняет хотя бы одну колонку, ПУСТУЮ в якорной строке.

        Последнее условие отсекает титульные строки прайса, чей текст лежит
        в уже занятой якорём колонке (иначе «услуги» из титула перебило бы
        настоящую колонку наименования)."""
        cells = [self._cell_text(c) for c in row]
        joined = " ".join(cells).lower()
        if not joined.strip():
            return False
        if not any(a in joined for a in HEADER_ANCHORS):
            return False
        # Цена в строке => это данные, не шапка.
        for cell in row:
            if parse_amount(cell) is not None:
                return False
        # Должна дополнять пустую в якоре колонку (а не дублировать титул).
        for idx, text in enumerate(cells):
            if text and (idx >= len(anchor) or not anchor[idx]):
                return True
        return False

    @classmethod
    def _merge_header_band(cls, band: list[list], anchor_offset: int) -> list[str]:
        """Сливает полосу строк шапки в одну подпись на колонку.

        За основу берём якорную строку (anchor_offset — её индекс в band).
        Остальные строки лишь ЗАПОЛНЯЮТ пустые в якоре колонки. Приоритет при
        конфликте: код > имя > цена > текст (двухстрочная шапка Клиники 8, где
        над колонкой кода через merge_cells висит общий заголовок «Цена...»).
        """
        width = max((len(r) for r in band), default=0)
        anchor_row = band[anchor_offset] if 0 <= anchor_offset < len(band) else []
        merged = [cls._cell_text(anchor_row[i]) if i < len(anchor_row) else ""
                  for i in range(width)]

        def rank(text: str) -> int:
            low = text.lower()
            if any(h in low for h in _CODE_CAPTION_HINTS):
                return 3
            if any(h in low for h in NAME_HINTS):
                return 2
            if any(h in low for h in _PRICE_CAPTION_HINTS):
                return 1
            return 0

        best_rank = [rank(c) for c in merged]
        for offset, row in enumerate(band):
            if offset == anchor_offset:
                continue
            for idx in range(len(row)):
                text = cls._cell_text(row[idx])
                if not text:
                    continue
                # Заполняем только пустые в якоре колонки; код перебивает цену.
                if not merged[idx]:
                    best_rank[idx], merged[idx] = rank(text), text
                elif rank(text) > best_rank[idx]:
                    best_rank[idx], merged[idx] = rank(text), text
        return merged

    # ------------------------------------------------------------------
    # Сопоставление колонок
    # ------------------------------------------------------------------
    def _map_columns(self, header_cells: list[str]) -> dict:
        """По ячейкам заголовка строит карту колонок.

        Возвращает словарь:
          name -> индекс колонки наименования (или None)
          code -> индекс колонки кода (или None)
          price_labels -> {индекс_колонки: исходная_подпись}
        """
        name_idx: int | None = None
        code_idx: int | None = None
        price_labels: dict[int, str] = {}

        for idx, cell in enumerate(header_cells):
            low = cell.lower().strip()
            if not low:
                continue
            # Колонка кода имеет приоритет над ценой: «Код по тарификатору»
            # не должен попасть в тарифы (ловушка Клиники 8).
            if any(h in low for h in _CODE_CAPTION_HINTS) and code_idx is None:
                code_idx = idx
            elif any(h in low for h in NAME_HINTS) and name_idx is None:
                name_idx = idx
            elif any(h in low for h in _PRICE_CAPTION_HINTS):
                price_labels[idx] = cell.strip()

        # Фолбэк: если колонку имени не нашли по якорю — самая длинная по
        # подписи текстовая колонка, не помеченная как код/цена (обычно это
        # «Наименование услуги»).
        if name_idx is None:
            best_len = -1
            for idx, cell in enumerate(header_cells):
                text = cell.strip()
                if not text or idx == code_idx or idx in price_labels:
                    continue
                if len(text) > best_len:
                    best_len, name_idx = len(text), idx

        return {"name": name_idx, "code": code_idx, "price_labels": price_labels}

    # ------------------------------------------------------------------
    # Утилиты работы с ячейками
    # ------------------------------------------------------------------
    @staticmethod
    def _cell_at(row: list, idx: int | None):
        if idx is None or idx >= len(row):
            return None
        return row[idx]

    @staticmethod
    def _cell_text(value) -> str:
        """Текст ячейки без хвостовых пробелов. None/пусто -> ''."""
        if value is None:
            return ""
        if isinstance(value, str):
            return value.strip()
        # Числа/даты приводим к строке (для кодов вроде «U1.1», номеров и т.п.).
        return str(value).strip()

    @classmethod
    def _first_text(cls, row: list) -> str:
        """Первая непустая текстовая ячейка строки (для съехавших секций)."""
        for value in row:
            text = cls._cell_text(value)
            if text and any(ch.isalpha() for ch in text):
                return text
        return ""
