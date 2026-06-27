"""Загрузчик справочника услуг (раздел 8.1, 3.3).

Гибкий маппинг полей (service_id, service_name, synonyms, category) из XLSX
или JSON. Эмбеддинг названия+синонимов считается один раз и кладётся в
Service.embedding (pgvector). До получения финального справочника работаем на
синтетическом, собранном из самих прайсов.
"""

from __future__ import annotations

import json
from collections import Counter

from sqlalchemy import select
from sqlalchemy.orm import Session

from app.models import PriceItem, Service
from app.normalization import embeddings as emb
from app.normalization.normalize import normalize

# Гибкий маппинг возможных названий колонок справочника.
# Реальный «Справочник услуг.xlsx»: Name_ru (название), Специальность (категория),
# TarificatrCode (код тарификатора -> кладём в icd_code).
FIELD_ALIASES = {
    "service_name": ("service_name", "name", "name_ru", "наименование", "услуга", "название"),
    "category": ("category", "категория", "раздел", "группа", "специальность"),
    "icd_code": ("icd_code", "icd", "мкб", "код", "tarificatrcode", "тарификатор", "код тарификатора"),
    "synonyms": ("synonyms", "синонимы", "synonym"),
}


def _pick(row: dict, field: str):
    for alias in FIELD_ALIASES[field]:
        for key in row:
            if str(key).strip().lower() == alias:
                return row[key]
    return None


def load_reference_records(path: str) -> list[dict]:
    """Читает справочник из JSON или XLSX в список словарей."""
    if path.lower().endswith(".json"):
        with open(path, encoding="utf-8") as fh:
            data = json.load(fh)
        return data if isinstance(data, list) else data.get("services", [])

    # XLSX
    from openpyxl import load_workbook

    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return []
    header = [str(c or "").strip() for c in rows[0]]
    records = []
    for r in rows[1:]:
        records.append({header[i]: r[i] for i in range(min(len(header), len(r)))})
    return records


def upsert_services(db: Session, records: list[dict], compute_embeddings: bool = True) -> int:
    """Создаёт/обновляет услуги справочника. Возвращает число добавленных.

    Дедуп по паре (нормализованное имя + код тарификатора): в справочнике
    одинаковые названия встречаются в разных специальностях с разными кодами,
    схлопывать их по одному имени нельзя (раздел 2, замечание о дедупликации).
    """
    added = 0
    seen: set[tuple[str, str | None]] = set()
    for rec in records:
        name = _pick(rec, "service_name")
        if not name:
            continue
        name = str(name).strip()
        norm = normalize(name)
        icd_raw = _pick(rec, "icd_code")
        icd = str(icd_raw).strip() if icd_raw not in (None, "") else None

        key = (norm, icd)
        if key in seen:
            continue
        existing = db.execute(
            select(Service).where(
                Service.service_name_normalized == norm,
                Service.icd_code == icd if icd is not None else Service.icd_code.is_(None),
            )
        ).scalar_one_or_none()
        if existing:
            seen.add(key)
            continue

        syn = _pick(rec, "synonyms") or []
        if isinstance(syn, str):
            syn = [s.strip() for s in syn.split(";") if s.strip()]
        category = _pick(rec, "category")
        svc = Service(
            service_name=name,
            service_name_normalized=norm,
            synonyms=list(syn),
            category=str(category).strip() if category else None,
            icd_code=icd,
        )
        db.add(svc)
        seen.add(key)
        added += 1
    db.flush()

    if compute_embeddings:
        embed_missing(db)
    db.commit()
    return added


def embed_missing(db: Session, batch_size: int = 128) -> int:
    """Считает эмбеддинги для услуг без embedding. Кэшируется, можно повторять."""
    services = db.execute(
        select(Service).where(Service.embedding.is_(None), Service.is_active.is_(True))
    ).scalars().all()
    if not services:
        return 0
    done = 0
    for i in range(0, len(services), batch_size):
        chunk = services[i : i + batch_size]
        texts = [
            (s.service_name + " " + " ".join(s.synonyms or [])).strip() for s in chunk
        ]
        vectors = emb.embed_batch(db, texts)
        for svc, vec in zip(chunk, vectors):
            if vec is not None:
                svc.embedding = vec
                done += 1
        db.flush()
    db.commit()
    return done


def build_synthetic_reference(db: Session, min_count: int = 1) -> int:
    """Синтетический справочник из самих прайсов (раздел 3.3).

    Группирует сырые названия позиций по нормализованной форме, самое частое
    написание становится каноническим, остальные — синонимами.
    """
    rows = db.execute(
        select(PriceItem.service_name_raw, PriceItem.category)
    ).all()
    groups: dict[str, Counter] = {}
    cat_of: dict[str, str | None] = {}
    for raw, category in rows:
        if not raw or not raw.strip():
            continue
        norm = normalize(raw)
        if not norm:
            continue
        groups.setdefault(norm, Counter())[raw.strip()] += 1
        if category and norm not in cat_of:
            cat_of[norm] = category

    added = 0
    for norm, variants in groups.items():
        if sum(variants.values()) < min_count:
            continue
        existing = db.execute(
            select(Service).where(Service.service_name_normalized == norm)
        ).scalar_one_or_none()
        if existing:
            continue
        canonical, _ = variants.most_common(1)[0]
        synonyms = [v for v in variants if v != canonical]
        db.add(
            Service(
                service_name=canonical,
                service_name_normalized=norm,
                synonyms=synonyms,
                category=cat_of.get(norm),
            )
        )
        added += 1
    db.commit()
    return added
